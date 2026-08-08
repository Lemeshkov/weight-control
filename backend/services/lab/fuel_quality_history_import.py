from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

import lab_models as lm

SNAPSHOT_FIELDS = ("wr_percent", "wa_percent", "aa_percent", "ar_percent", "ad_percent", "va_percent",
    "vdaf_percent", "vr_percent", "sa_percent", "sr_percent", "sd_percent", "qi_r_kcal_kg")


class LegacyImportError(ValueError):
    pass


@dataclass(frozen=True)
class LegacyFuelQualityRow:
    excel_row: int
    sample_date: date
    values: dict[str, Decimal]
    import_key: str


@dataclass(frozen=True)
class LegacyImportResult:
    found: int
    existing: int
    imported: int
    dry_run: bool


def _decimal(value, row: int, column: str) -> Decimal:
    if value is None or isinstance(value, bool):
        raise LegacyImportError(f"Строка {row}, столбец {column}: требуется числовое значение")
    try:
        result = Decimal(str(value))
    except (InvalidOperation, ValueError) as exc:
        raise LegacyImportError(f"Строка {row}, столбец {column}: некорректное число {value!r}") from exc
    if not result.is_finite():
        raise LegacyImportError(f"Строка {row}, столбец {column}: число должно быть конечным")
    return result


def read_legacy_rows(workbook_path: Path, year: int, month: int) -> list[LegacyFuelQualityRow]:
    path = Path(workbook_path)
    if not path.is_file():
        raise LegacyImportError(f"Файл не найден: {path}")
    sheet_name = f"{month:02d}"
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise LegacyImportError(f"Лист {sheet_name!r} не найден")
        sheet = workbook[sheet_name]
        result = []
        for excel_row in range(3, sheet.max_row + 1):
            raw = [sheet.cell(excel_row, column).value for column in range(1, 14)]
            if all(value is None for value in raw):
                continue
            raw_date = raw[0]
            if isinstance(raw_date, datetime):
                raw_date = raw_date.date()
            if not isinstance(raw_date, date) or raw_date.year != year or raw_date.month != month:
                raise LegacyImportError(f"Строка {excel_row}, столбец A: дата должна относиться к {sheet_name}.{year}")
            values = {name: _decimal(value, excel_row, chr(66 + offset)) for offset, (name, value) in enumerate(zip(SNAPSHOT_FIELDS, raw[1:]))}
            key = f"fuel-quality:{year}:{sheet_name}:{raw_date.isoformat()}"
            result.append(LegacyFuelQualityRow(excel_row, raw_date, values, key))
        if not result:
            raise LegacyImportError(f"На листе {sheet_name!r} нет строк данных")
        dates = [row.sample_date for row in result]
        if len(dates) != len(set(dates)):
            raise LegacyImportError(f"На листе {sheet_name!r} обнаружены повторяющиеся даты")
        return result
    finally:
        workbook.close()


def import_legacy_rows(db: Session, workbook_path: Path, year: int, month: int, apply: bool = False) -> LegacyImportResult:
    rows = read_legacy_rows(workbook_path, year, month)
    keys = [row.import_key for row in rows]
    existing = set(value for value, in db.query(lm.LabFuelQualityTest.legacy_import_key).filter(
        lm.LabFuelQualityTest.legacy_import_key.in_(keys)).all())
    pending = [row for row in rows if row.import_key not in existing]
    if not apply:
        return LegacyImportResult(len(rows), len(existing), 0, True)
    try:
        for row in pending:
            item = lm.LabFuelQualityTest(
                sample_date=row.sample_date,
                sample_name=f"Ежесуточный контроль {row.sample_date:%d.%m.%Y}",
                calorimeter=None, sa_percent=row.values["sa_percent"], alpha=None,
                wa_percent=row.values["wa_percent"], aa_percent=row.values["aa_percent"],
                wr_percent=row.values["wr_percent"], hydrogen_input_percent=None,
                qb_a_1_kcal_kg=None, qb_a_2_kcal_kg=None, va_percent=row.values["va_percent"],
                status=lm.LabExperimentStatus.COMPLETED, lab_technician_name=None,
                calculation_snapshot={key: str(value) for key, value in row.values.items()},
                source="LEGACY_EXCEL", source_file=Path(workbook_path).name, source_sheet=f"{month:02d}",
                legacy_import_key=row.import_key,
            )
            db.add(item)
            db.flush()
            db.add(lm.LabFuelQualityAuditLog(test_id=item.id, action="IMPORT_EXCEL", new_values={
                "source_file": item.source_file, "source_sheet": item.source_sheet,
                "legacy_import_key": item.legacy_import_key, "excel_row": row.excel_row,
            }))
        db.commit()
    except Exception:
        db.rollback()
        raise
    return LegacyImportResult(len(rows), len(existing), len(pending), False)
