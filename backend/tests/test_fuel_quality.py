from datetime import date
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pytest
import models  # noqa: F401 - register shared users table in SQLAlchemy metadata
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import Workbook,load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from database import Base,get_db
import lab_models as lm
from routers.laboratory import router
from services.lab.fuel_quality_calculations import calculate_fuel_quality
from services.lab.fuel_quality_history_import import LegacyImportError, import_legacy_rows

PDF_INPUT={"sa_percent":"0.37","alpha":"0.0015","wa_percent":"2.06","aa_percent":"11.72","wr_percent":"11.99",
    "hydrogen_input_percent":"5.56","qb_a_1_kcal_kg":"6923.00","qb_a_2_kcal_kg":"6924.00","va_percent":"33.88"}
PAYLOAD={"sample_date":"2026-07-01","sample_name":"Лента 01.07.2026","calorimeter":"C200",
    **PDF_INPUT,"lab_technician_name":"Шкапорова С.Л.","wagon_numbers":"1, 2","invoice_number":"TN-1"}


@pytest.fixture()
def client():
    engine=create_engine("sqlite://",connect_args={"check_same_thread":False},poolclass=StaticPool)
    Base.metadata.create_all(engine);Session=sessionmaker(bind=engine);app=FastAPI();app.include_router(router)
    def override():
        db=Session()
        try:yield db
        finally:db.close()
    app.dependency_overrides[get_db]=override
    yield TestClient(app),Session
    Base.metadata.drop_all(engine)


def test_pdf_regression_fixture():
    values=calculate_fuel_quality(**{key:Decimal(value) for key,value in PDF_INPUT.items()})
    expected={"qb_a_kcal_kg":"6923.50","qs_a_kcal_kg":"6904.79","ad_percent":"11.97","ar_percent":"10.53",
        "vdaf_percent":"39.29","vr_percent":"30.44","sr_percent":"0.33","sd_percent":"0.38",
        "qs_r_kcal_kg":"6204.72","qi_a_kcal_kg":"6602.99","hr_percent":"4.31","qi_r_kcal_kg":"5910.18",
        "qs_daf_kcal_kg":"8008.34","wmax_daf_percent":"13.40","qs_af_kcal_kg":"6935.11","qb_daf_kcal_kg":"8030.04"}
    assert values=={key:Decimal(value) for key,value in expected.items()}


def test_invalid_denominator():
    values={key:Decimal(value) for key,value in PDF_INPUT.items()};values.update(wa_percent=Decimal("60"),aa_percent=Decimal("40"))
    with pytest.raises(ValueError,match="знаменатель"):calculate_fuel_quality(**values)


def test_draft_update_complete_immutable_archive_audit_and_filters(client):
    api,Session=client
    created=api.post("/api/v1/laboratory/fuel-quality",json=PAYLOAD)
    assert created.status_code==201 and created.json()["status"]=="DRAFT"
    test_id=created.json()["id"];stamp=created.json()["updated_at"]
    updated=api.put(f"/api/v1/laboratory/fuel-quality/{test_id}",json={**PAYLOAD,"sample_name":"Обновлено","expected_updated_at":stamp})
    assert updated.status_code==200
    completed=api.post(f"/api/v1/laboratory/fuel-quality/{test_id}/complete")
    assert completed.status_code==200 and completed.json()["status"]=="COMPLETED"
    assert api.put(f"/api/v1/laboratory/fuel-quality/{test_id}",json={**PAYLOAD,"expected_updated_at":completed.json()["updated_at"]}).status_code==409
    listing=api.get("/api/v1/laboratory/fuel-quality",params={"year":2026,"month":7,"status":"COMPLETED","search":"Обновлено","limit":1})
    assert listing.status_code==200 and listing.json()["total"]==1
    archived=api.post(f"/api/v1/laboratory/fuel-quality/{test_id}/archive")
    assert archived.status_code==200 and archived.json()["status"]=="ARCHIVED"
    audit=api.get(f"/api/v1/laboratory/fuel-quality/{test_id}/audit-log").json()
    assert [x["action"] for x in audit]==["CREATE","UPDATE","COMPLETE","ARCHIVE"]


@pytest.mark.parametrize("month,days",[(1,31),(2,28),(4,30)])
def test_excel_export_valid_month_lengths_and_mapping(client,month,days):
    api,_=client
    if month==1:
        payload={**PAYLOAD,"sample_date":"2026-01-01"};created=api.post("/api/v1/laboratory/fuel-quality",json=payload).json();api.post(f"/api/v1/laboratory/fuel-quality/{created['id']}/complete")
    response=api.get("/api/v1/laboratory/fuel-quality/export.xlsx",params={"year":2026,"month":month})
    assert response.status_code==200 and response.content[:2]==b"PK"
    workbook=load_workbook(BytesIO(response.content));sheet=workbook[f"{month:02d}"]
    assert workbook.sheetnames==[f"{month:02d}"] and sheet.max_column==13 and sheet.cell(days+2,1).value.date()==date(2026,month,days)
    assert "A1:M1" in {str(x) for x in sheet.merged_cells.ranges}
    assert sheet.column_dimensions["A"].width and sheet["B2"].alignment.wrap_text
    if month==1: assert [sheet.cell(3,c).value for c in range(2,14)]==[11.99,2.06,11.72,10.53,11.97,33.88,39.29,30.44,0.37,0.33,0.38,5910]


def test_historical_july_import_dry_run_apply_idempotency_and_export(client):
    api,Session=client
    source=Path(__file__).resolve().parents[2]/"docs"/"reference"/"Ежесуточный контроль топлива 2026.xlsx"
    db=Session()
    preview=import_legacy_rows(db,source,2026,7)
    assert (preview.found,preview.existing,preview.imported,preview.dry_run)==(31,0,0,True)
    assert db.query(lm.LabFuelQualityTest).count()==0
    applied=import_legacy_rows(db,source,2026,7,apply=True)
    assert (applied.found,applied.existing,applied.imported)==(31,0,31)
    repeated=import_legacy_rows(db,source,2026,7,apply=True)
    assert (repeated.existing,repeated.imported)==(31,0)
    rows=db.query(lm.LabFuelQualityTest).order_by(lm.LabFuelQualityTest.sample_date).all()
    assert [row.sample_date for row in rows]==[date(2026,7,day) for day in range(1,32)]
    first=rows[0]
    assert first.source=="LEGACY_EXCEL" and first.status==lm.LabExperimentStatus.COMPLETED
    assert first.alpha is None and first.hydrogen_input_percent is None and first.qb_a_1_kcal_kg is None
    assert first.calculation_snapshot["qi_r_kcal_kg"]=="5910"
    assert db.query(lm.LabFuelQualityAuditLog).filter_by(action="IMPORT_EXCEL").count()==31
    db.close()
    listing=api.get("/api/v1/laboratory/fuel-quality",params={"year":2026,"month":7,"limit":200}).json()
    assert listing["total"]==31 and all(item["source"]=="LEGACY_EXCEL" for item in listing["items"])
    assert api.put(f"/api/v1/laboratory/fuel-quality/{first.id}",json=PAYLOAD).status_code==409
    assert api.post(f"/api/v1/laboratory/fuel-quality/{first.id}/archive").status_code==409
    exported=load_workbook(BytesIO(api.get("/api/v1/laboratory/fuel-quality/export.xlsx",params={"year":2026,"month":7}).content),data_only=True)["07"]
    original=load_workbook(source,read_only=True,data_only=True)["07"]
    for row in range(3,34):
        assert [exported.cell(row,column).value for column in range(1,14)]==[original.cell(row,column).value for column in range(1,14)]


def test_historical_import_invalid_row_rolls_back(client,tmp_path):
    _,Session=client
    path=tmp_path/"invalid.xlsx";workbook=Workbook();sheet=workbook.active;sheet.title="07"
    sheet.append([]);sheet.append([])
    sheet.append([date(2026,7,1),11.99,2.06,11.72,10.53,11.97,33.88,39.29,30.44,.37,.33,.38,5910])
    sheet.append([date(2026,7,2),12,"bad",11,10,12,34,40,30,.4,.3,.4,5900]);workbook.save(path)
    db=Session()
    with pytest.raises(LegacyImportError,match="столбец C"):
        import_legacy_rows(db,path,2026,7,apply=True)
    assert db.query(lm.LabFuelQualityTest).count()==0


def test_manual_api_stays_strict_and_defaults_source(client):
    api,_=client
    invalid={key:value for key,value in PAYLOAD.items() if key!="alpha"}
    assert api.post("/api/v1/laboratory/fuel-quality",json=invalid).status_code==422
    created=api.post("/api/v1/laboratory/fuel-quality",json=PAYLOAD)
    assert created.status_code==201 and created.json()["source"]=="MANUAL"
