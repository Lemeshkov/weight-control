from datetime import date, datetime, timedelta
from decimal import Decimal
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import case, func, or_
from sqlalchemy.orm import Session, joinedload

from config import settings
from database import get_db
from lab_models import CoalGrade, Supplier
from models import (CoalAcceptance, CoalAcceptanceAuditLog, CoalAcceptanceStatus,
                    LidarPassSession, Trip, Vehicle)
from schemas.coal_acceptance import AcceptanceUpdate, AcceptanceWrite, CompleteRequest
from services.coal_acceptance import actual_net_weight_t, calculate, contract_date, localized_moments

router = APIRouter(prefix="/api/coal-acceptance", tags=["coal-acceptance"])
FIELDS = ("shipment_date", "act_number", "transport_invoice_number", "document_net_weight_t", "supplier_id", "coal_grade_id", "uk_number", "invoice_number", "receiver_name", "notes")


def _plain(a):
    if not a: return None
    return {k: (str(v) if isinstance(v, (Decimal, date, datetime)) else v) for k, v in {f: getattr(a, f) for f in FIELDS}.items()}


def _latest_lidar(trip):
    return max(trip.lidar_pass_sessions, key=lambda x: (x.started_at, x.id), default=None)


def _item(trip):
    a, lidar = trip.coal_acceptance, _latest_lidar(trip)
    actual = actual_net_weight_t(trip)
    values = calculate(actual, a.document_net_weight_t if a else None, settings.COAL_ACCEPTANCE_WEIGHT_TOLERANCE)
    local, moscow = localized_moments(trip, settings.COAL_ACCEPTANCE_LOCAL_TIMEZONE)
    status = a.status.value if a else "WAITING"
    warning = actual is None or (values["shortage_t"] or Decimal(0)) > 0
    return {
        "trip_id": trip.id, "status": status, "has_warning": warning,
        "vehicle": {"license_plate": trip.vehicle.plate_number if trip.vehicle else None, "model": trip.vehicle.model if trip.vehicle else None},
        "entry_time": trip.entry_time, "exit_time": trip.exit_time,
        "acceptance_time_local": local, "acceptance_time_moscow": moscow,
        "contract_date": contract_date(local), "actual_net_weight_t": actual,
        "brutto_weight_kg": None if not trip.entry_measurement else trip.entry_measurement.weight_brutto,
        "tare_weight_kg": None if not trip.exit_measurement else trip.exit_measurement.weight_tare,
        "weight_stable": bool(lidar and lidar.stable_weight_at),
        "lidar": None if not lidar else {"status": lidar.status, "profiles_count": lidar.profiles_count, "volume_status": lidar.volume_status, "estimated_volume_m3": lidar.estimated_volume_m3},
        "acceptance": None if not a else {**_plain(a), "id": a.id, "status": a.status.value, "updated_at": a.updated_at,
            "supplier": None if not a.supplier else {"id": a.supplier.id, "name": a.supplier.name},
            "coal_grade": None if not a.coal_grade else {"id": a.coal_grade.id, "name": a.coal_grade.name}},
        "calculated": values,
    }


def _query(db):
    return db.query(Trip).options(joinedload(Trip.vehicle), joinedload(Trip.entry_measurement), joinedload(Trip.exit_measurement), joinedload(Trip.lidar_pass_sessions), joinedload(Trip.coal_acceptance).joinedload(CoalAcceptance.supplier), joinedload(Trip.coal_acceptance).joinedload(CoalAcceptance.coal_grade))


def _get(db, trip_id):
    trip = _query(db).filter(Trip.id == trip_id).one_or_none()
    if not trip: raise HTTPException(404, "Trip not found")
    return trip


def _validate_refs(db, payload):
    if payload.supplier_id and not db.query(Supplier).filter(Supplier.id == payload.supplier_id, Supplier.is_active.is_(True)).first(): raise HTTPException(422, "Unknown supplier")
    if payload.coal_grade_id and not db.query(CoalGrade).filter(CoalGrade.id == payload.coal_grade_id, CoalGrade.is_active.is_(True)).first(): raise HTTPException(422, "Unknown coal grade")


def _audit(db, a, action, before, operator):
    db.add(CoalAcceptanceAuditLog(coal_acceptance_id=a.id, trip_id=a.trip_id, action=action, changed_by_name=operator, previous_values=before, new_values=_plain(a)))


@router.get("/queue")
def queue(db: Session = Depends(get_db), page: int = Query(1, ge=1), page_size: int = Query(25, ge=1, le=100), date_from: date | None = None, date_to: date | None = None, status: str | None = None, plate: str | None = None, invoice_number: str | None = None, supplier_id: int | None = None, coal_grade_id: int | None = None):
    today = date.today(); date_from = date_from or today; date_to = date_to or today
    q = _query(db).outerjoin(CoalAcceptance).outerjoin(Vehicle).filter(Trip.entry_time >= datetime.combine(date_from, datetime.min.time()), Trip.entry_time < datetime.combine(date_to + timedelta(days=1), datetime.min.time()))
    if status == "WAITING": q = q.filter(CoalAcceptance.id.is_(None))
    elif status in ("DRAFT", "COMPLETED"): q = q.filter(CoalAcceptance.status == CoalAcceptanceStatus(status))
    if plate: q = q.filter(Vehicle.plate_number.ilike(f"%{plate}%"))
    if invoice_number: q = q.filter(CoalAcceptance.transport_invoice_number.ilike(f"%{invoice_number}%"))
    if supplier_id: q = q.filter(CoalAcceptance.supplier_id == supplier_id)
    if coal_grade_id: q = q.filter(CoalAcceptance.coal_grade_id == coal_grade_id)
    status_order = case(
        (CoalAcceptance.id.is_(None), 0),
        (CoalAcceptance.status == CoalAcceptanceStatus.DRAFT, 1),
        else_=2,
    )
    total = q.order_by(None).count(); trips = q.order_by(status_order, Trip.entry_time.desc()).offset((page-1)*page_size).limit(page_size).all()
    return {"items": [_item(x) for x in trips], "page": page, "page_size": page_size, "total": total}


@router.get("/directories")
def directories(db: Session = Depends(get_db)):
    return {"suppliers": [{"id": x.id, "code": x.code, "name": x.name} for x in db.query(Supplier).filter(Supplier.is_active.is_(True)).order_by(Supplier.name)], "coal_grades": [{"id": x.id, "code": x.code, "name": x.name} for x in db.query(CoalGrade).filter(CoalGrade.is_active.is_(True)).order_by(CoalGrade.name)]}


@router.get("/export.xlsx")
def export_xlsx(db: Session = Depends(get_db), date_from: date | None = None, date_to: date | None = None):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError: raise HTTPException(503, "Excel export dependency is unavailable")
    q = _query(db).join(CoalAcceptance).filter(CoalAcceptance.status == CoalAcceptanceStatus.COMPLETED)
    if date_from: q = q.filter(Trip.entry_time >= datetime.combine(date_from, datetime.min.time()))
    if date_to: q = q.filter(Trip.entry_time < datetime.combine(date_to + timedelta(days=1), datetime.min.time()))
    trips = q.order_by(Trip.entry_time, Trip.id).all()
    headers = ["№ п/п", "Дата отгрузки (транспортной накладной)", "№ АКТа", "Номер транспортной накладной", "Масса груза по ТН (нетто), тн.", "Фактически взвешанный вес (нетто), тн", "Масса оприходованного топлива (на склад), т", "Расхождение (Ф-ТН), тн.", "Допустимое предельное расхождение", "Норма естественной убыли", "Недостача массы", "Излишки массы", "Масса топлива, подлежащего списанию", "Грузоотправитель", "Марка угля", "Номер УК", "Госномер автомобиля", "Дата, время приёмки по местному времени", "Дата, время приёмки по МСК", "ФИО приёмосдатчика", "№ СФ", "Дата по договору", "Принято за сутки по договору"]
    wb=Workbook(); ws=wb.active; ws.title="Приёмка угля"; ws.append(headers)
    for c in ws[1]: c.font=Font(bold=True,color="FFFFFF"); c.fill=PatternFill("solid",fgColor="1F4E78"); c.alignment=Alignment(wrap_text=True,vertical="center")
    for n, trip in enumerate(trips, 1):
        x=_item(trip); a=trip.coal_acceptance; row=n+1
        ws.append([n,a.shipment_date,a.act_number,a.transport_invoice_number,float(a.document_net_weight_t),float(x["actual_net_weight_t"]) if x["actual_net_weight_t"] is not None else None,f"=E{row}-K{row}-M{row}+L{row}",f"=F{row}-E{row}",f"=ROUND(E{row}*{settings.COAL_ACCEPTANCE_WEIGHT_TOLERANCE},3)",0,f'=IF(H{row}>0,0,IF(ABS(H{row})<I{row},0,ABS(H{row})))',f'=IF(H{row}<0,0,IF(ABS(H{row})>I{row},ABS(H{row}),0))',0,a.supplier.name,a.coal_grade.name,a.uk_number,x["vehicle"]["license_plate"],x["acceptance_time_local"].replace(tzinfo=None),x["acceptance_time_moscow"].replace(tzinfo=None),a.receiver_name,a.invoice_number,x["contract_date"],f'=SUMIF(V:V,V{row},G:G)'])
    ws.freeze_panes="A2"; ws.auto_filter.ref=f"A1:W{max(1,ws.max_row)}"; ws.row_dimensions[1].height=55
    for col in range(1,24): ws.column_dimensions[__import__('openpyxl').utils.get_column_letter(col)].width=18
    for row in ws.iter_rows(min_row=2,min_col=5,max_col=13):
        for cell in row: cell.number_format='0.000'
    output=BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": 'attachment; filename="coal_acceptance.xlsx"'})


@router.get("/{trip_id}")
def get_acceptance(trip_id: int, db: Session = Depends(get_db)): return _item(_get(db, trip_id))


@router.post("/{trip_id}", status_code=201)
def create_acceptance(trip_id: int, payload: AcceptanceWrite, db: Session = Depends(get_db)):
    trip=_get(db,trip_id)
    if trip.coal_acceptance: raise HTTPException(409,"Acceptance already exists")
    _validate_refs(db,payload); data=payload.model_dump(exclude={"operator_name"}); a=CoalAcceptance(trip_id=trip_id,**data); db.add(a); db.flush(); _audit(db,a,"CREATE",None,payload.operator_name); db.commit(); return _item(_get(db,trip_id))


@router.put("/{trip_id}")
def update_acceptance(trip_id: int, payload: AcceptanceUpdate, db: Session = Depends(get_db)):
    trip=_get(db,trip_id); a=trip.coal_acceptance
    if not a: raise HTTPException(404,"Acceptance not found")
    if a.status == CoalAcceptanceStatus.COMPLETED: raise HTTPException(409,"Completed acceptance is read-only")
    expected=payload.expected_updated_at
    current=a.updated_at
    if current.tzinfo is None and expected.tzinfo is not None: expected=expected.replace(tzinfo=None)
    if current != expected: raise HTTPException(409,"Acceptance was changed by another operator")
    _validate_refs(db,payload); before=_plain(a)
    for key,value in payload.model_dump(exclude={"operator_name","expected_updated_at"}).items(): setattr(a,key,value)
    a.updated_at=datetime.now().astimezone(); _audit(db,a,"UPDATE",before,payload.operator_name); db.commit(); return _item(_get(db,trip_id))


@router.post("/{trip_id}/complete")
def complete_acceptance(trip_id: int, payload: CompleteRequest, db: Session = Depends(get_db)):
    trip=_get(db,trip_id); a=trip.coal_acceptance
    if not a: raise HTTPException(404,"Acceptance not found")
    if a.status == CoalAcceptanceStatus.COMPLETED: return _item(trip)
    expected=payload.expected_updated_at
    current=a.updated_at
    if current.tzinfo is None and expected.tzinfo is not None: expected=expected.replace(tzinfo=None)
    if current != expected: raise HTTPException(409,"Acceptance was changed by another operator")
    required=("shipment_date","act_number","transport_invoice_number","document_net_weight_t","supplier_id","coal_grade_id","receiver_name")
    missing=[x for x in required if getattr(a,x) in (None,"")]
    if actual_net_weight_t(trip) is None: missing.append("actual_net_weight_t")
    if missing: raise HTTPException(422,{"missing_fields":missing})
    before=_plain(a); a.status=CoalAcceptanceStatus.COMPLETED; a.updated_at=datetime.now().astimezone(); _audit(db,a,"COMPLETE",before,payload.operator_name); db.commit(); return _item(_get(db,trip_id))

