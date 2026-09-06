from datetime import date
from io import BytesIO
from urllib.parse import quote

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy.orm import Session

from database import get_db
from schemas.analytics import SupplierSummaryResponse
from services.analytics.supplier_summary import build_supplier_summary

router=APIRouter(prefix="/api/analytics",tags=["analytics"])

def _report(date_from:date,date_to:date,supplier_id:int|None,db:Session):
    if date_from>date_to:raise HTTPException(422,"Дата начала не может быть позже даты окончания")
    return build_supplier_summary(db,date_from,date_to,supplier_id)

@router.get("/supplier-summary",response_model=SupplierSummaryResponse)
def supplier_summary(date_from:date,date_to:date,supplier_id:int|None=None,db:Session=Depends(get_db)):
    return _report(date_from,date_to,supplier_id,db)

@router.get("/supplier-summary/export")
def supplier_summary_export(date_from:date,date_to:date,supplier_id:int|None=None,db:Session=Depends(get_db)):
    report=_report(date_from,date_to,supplier_id,db);book=Workbook();sheet=book.active;sheet.title="По поставщикам"
    sheet.append(["Сводка по поставщикам"]);sheet.append([f"Период: {date_from.isoformat()} — {date_to.isoformat()}"])
    if supplier_id is not None:sheet.append([f"Поставщик: {report.rows[0].supplier_name if report.rows else supplier_id}"])
    header=sheet.max_row+2;sheet.append(["Поставщик","Количество рейсов","Вес, т","Объем, м³","Насыпная плотность, т/м³"])
    for cell in sheet[header]:cell.font=Font(bold=True)
    for row in report.rows:sheet.append([row.supplier_name,row.trip_count,float(row.total_weight_t),None,None])
    sheet.append(["ИТОГО",report.totals.trip_count,float(report.totals.total_weight_t),None,None]);sheet[sheet.max_row][0].font=Font(bold=True)
    sheet.freeze_panes=f"A{header+1}";sheet.column_dimensions["A"].width=35;sheet.column_dimensions["B"].width=20;sheet.column_dimensions["C"].width=16;sheet.column_dimensions["D"].width=16;sheet.column_dimensions["E"].width=30
    stream=BytesIO();book.save(stream);stream.seek(0);filename=f"supplier_summary_{date_from}_{date_to}.xlsx"
    return StreamingResponse(stream,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":f"attachment; filename*=UTF-8''{quote(filename)}"})
